import csv
import io

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from backend import db
from backend.api.leads import _build_filters

router = APIRouter(prefix="/api", tags=["export"])

COLUMNS = [
    "id", "place_id", "name", "category", "phone_raw", "phone_e164",
    "is_mobile_phone", "address", "city", "state", "postal_code",
    "latitude", "longitude", "google_maps_url", "rating", "reviews_count",
    "website_url", "final_url", "site_status", "https", "response_time_ms",
    "page_size_bytes", "has_title", "has_viewport", "has_contact_form",
    "site_tech_issues", "email", "instagram", "facebook", "linkedin",
    "whatsapp_found", "phone_on_site", "score", "score_class", "score_reasons",
    "crm_status", "notes", "first_seen_at", "updated_at",
]


def _fetch_rows(search_id, score_class, site_status, has_phone, has_whatsapp,
                 city, state, crm_status, min_reviews, min_score, max_score, q):
    where, params = _build_filters(
        search_id, score_class, site_status, has_phone, has_whatsapp,
        city, state, crm_status, min_reviews, min_score, max_score, q,
    )
    conn = db.get_conn()
    cur = conn.cursor()
    cur.execute(f"SELECT * FROM leads l WHERE {where} ORDER BY l.score DESC", params)
    return [dict(r) for r in cur.fetchall()]


@router.get("/export")
async def export_leads(
    format: str = "csv",
    search_id: int | None = None,
    score_class: str | None = None,
    site_status: str | None = None,
    has_phone: bool = False,
    has_whatsapp: bool = False,
    city: str | None = None,
    state: str | None = None,
    crm_status: str | None = None,
    min_reviews: int | None = None,
    min_score: int | None = None,
    max_score: int | None = None,
    q: str | None = None,
):
    rows = _fetch_rows(
        search_id, score_class, site_status, has_phone, has_whatsapp,
        city, state, crm_status, min_reviews, min_score, max_score, q,
    )

    if format == "xlsx":
        return _export_xlsx(rows)
    return _export_csv(rows)


def _export_csv(rows):
    buf = io.StringIO()
    buf.write("﻿")
    writer = csv.DictWriter(buf, fieldnames=COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=prospector_leads.csv"},
    )


def _export_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(COLUMNS)
    header_fill = PatternFill(start_color="111719", end_color="111719", fill_type="solid")
    header_font = Font(color="2BFF88", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        ws.append([row.get(col) for col in COLUMNS])
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prospector_leads.xlsx"},
    )
